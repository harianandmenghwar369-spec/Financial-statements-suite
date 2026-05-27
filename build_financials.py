from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
import copy

# ─── COLOR PALETTE ───────────────────────────────────────────────────────────
BLACK       = "FF000000"
WHITE       = "FFFFFFFF"
DARK_BG     = "FF0D0D0D"
GREEN       = "FF1DBF73"
GREEN_LIGHT = "FFE8FAF2"
GREEN_MID   = "FF0F6E56"
RED         = "FFDC2626"
RED_LIGHT   = "FFFEF2F2"
AMBER       = "FFD97706"
AMBER_LIGHT = "FFFEF3C7"
BLUE        = "FF1D4ED8"
BLUE_LIGHT  = "FFEFF6FF"
GRAY_DARK   = "FF374151"
GRAY_MID    = "FF6B7280"
GRAY_LIGHT  = "FFF9FAFB"
GRAY_BORDER = "FFE5E7EB"
INPUT_BLUE  = "FF1E3A8A"   # blue text = user input cells
FORMULA_BLK = "FF111827"  # black text = formula cells

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def thin(color=GRAY_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=GRAY_BORDER):
    s = Side(style='thin', color=color)
    return Border(bottom=s)

def thick_bottom(color="FF1DBF73"):
    s = Side(style='medium', color=color)
    return Border(bottom=s)

def font(bold=False, size=11, color=FORMULA_BLK, italic=False, name="Arial"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value, bold=False, size=11, color=FORMULA_BLK,
             bg=None, h_align="left", italic=False, border=None, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=size, color=color, italic=italic)
    if bg:
        c.fill = fill(bg)
    c.alignment = align(h=h_align)
    if border:
        c.border = border
    if num_fmt:
        c.number_format = num_fmt
    return c

def input_cell(ws, row, col, value, bg=None, num_fmt='$#,##0;($#,##0);"-"'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=False, size=11, color=INPUT_BLUE)
    if bg:
        c.fill = fill(bg)
    c.alignment = align(h="right")
    c.number_format = num_fmt
    c.border = thin()
    return c

def formula_cell(ws, row, col, formula, bold=False, color=FORMULA_BLK,
                 bg=None, num_fmt='$#,##0;($#,##0);"-"', border=None, size=11):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = font(bold=bold, size=size, color=color)
    if bg:
        c.fill = fill(bg)
    c.alignment = align(h="right")
    c.number_format = num_fmt
    if border:
        c.border = border
    return c

def pct_formula(ws, row, col, formula, bold=False, color=FORMULA_BLK, bg=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = font(bold=bold, size=11, color=color)
    if bg:
        c.fill = fill(bg)
    c.alignment = align(h="right")
    c.number_format = '0.0%;(0.0%);"-"'
    return c

def section_header(ws, row, label, col_start=2, col_end=5, bg=DARK_BG):
    ws.row_dimensions[row].height = 22
    c = ws.cell(row=row, column=col_start, value=label)
    c.font = font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill = fill(bg)
    c.alignment = align(h="left")
    for col in range(col_start, col_end+1):
        ws.cell(row=row, column=col).fill = fill(bg)

def cover_header(ws, row, label, col_start=1, col_end=6):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=label)
    c.font = font(bold=True, size=18, color=WHITE, name="Arial")
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")
    ws.row_dimensions[row].height = 36

def sub_header(ws, row, label, col_start=1, col_end=6, bg=GREEN, txt=WHITE):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=label)
    c.font = font(bold=True, size=11, color=txt, name="Arial")
    c.fill = fill(bg)
    c.alignment = align(h="center")
    ws.row_dimensions[row].height = 20

def label_row(ws, row, label, col=2, indent=False):
    prefix = "    " if indent else ""
    c = ws.cell(row=row, column=col, value=prefix + label)
    c.font = font(size=11, color=GRAY_DARK)
    c.alignment = align(h="left")
    c.border = bottom_border()
    ws.row_dimensions[row].height = 18

def total_row(ws, row, label, formula, col_label=2, col_val=5,
              bg=GREEN_LIGHT, txt_color=GREEN_MID, bold=True):
    ws.row_dimensions[row].height = 20
    c = ws.cell(row=row, column=col_label, value=label)
    c.font = font(bold=True, size=11, color=txt_color, name="Arial")
    c.fill = fill(bg)
    c.alignment = align(h="left")
    for col in range(col_label, col_val):
        ws.cell(row=row, column=col).fill = fill(bg)
    fc = formula_cell(ws, row, col_val, formula, bold=True,
                      color=txt_color, bg=bg,
                      border=thick_bottom(GREEN))
    return fc

def set_col_widths(ws, widths):
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(size=9, color=GRAY_MID, italic=True)
    c.alignment = align(h="left", wrap=True)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — INCOME STATEMENT
# ═══════════════════════════════════════════════════════════════════════════
def build_income_statement(wb):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [1, 34, 4, 4, 16, 16, 4])

    # ── Title block ──────────────────────────────────────────────────────
    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value = "INCOME STATEMENT"
    c.font = font(bold=True, size=20, color=WHITE, name="Arial")
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "The Solo Operator's Financial Toolkit  •  harianand07"
    c.font = font(size=10, color=GRAY_MID, italic=True)
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")
    ws.row_dimensions[2].height = 18

    # Column headers
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 22
    for col, label in [(2,"Line Item"), (5,"Amount ($)"), (6,"% of Revenue")]:
        c = ws.cell(row=4, column=col, value=label)
        c.font = font(bold=True, size=10, color=WHITE)
        c.fill = fill(GREEN_MID)
        c.alignment = align(h="center" if col>2 else "left")

    ws.row_dimensions[5].height = 6

    # ── REVENUE ──────────────────────────────────────────────────────────
    r = 6
    section_header(ws, r, "  REVENUE", 2, 6)
    r += 1

    rev_rows = []
    items = [
        ("Product / service sales", 45000),
        ("Consulting / project fees", 18000),
        ("Recurring / retainer income", 8000),
        ("Other income", 2500),
    ]
    for label, val in items:
        label_row(ws, r, label)
        input_cell(ws, r, 5, val)
        pct_formula(ws, r, 6, f"=IF(E{r+6}=0,0,E{r}/E{r+6})", color=GRAY_MID)
        rev_rows.append(r)
        r += 1

    disc_row = r
    label_row(ws, r, "Less: discounts & returns")
    input_cell(ws, r, 5, 1200)
    r += 1

    gross_rev_row = r
    total_row(ws, r, "Gross Revenue",
              f"=SUM(E{rev_rows[0]}:E{rev_rows[-1]})", col_val=5)
    formula_cell(ws, r, 6, "=1", bold=True, color=GREEN_MID, bg=GREEN_LIGHT,
                 num_fmt='0.0%')
    r += 1

    net_rev_row = r
    total_row(ws, r, "Net Revenue",
              f"=E{gross_rev_row}-E{disc_row}", col_val=5)
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r-1}/E{net_rev_row})",
                bold=True, color=GREEN_MID, bg=GREEN_LIGHT)
    r += 2

    # ── COGS ─────────────────────────────────────────────────────────────
    section_header(ws, r, "  COST OF GOODS SOLD (COGS)", 2, 6)
    r += 1
    cogs_rows = []
    for label, val in [("Raw materials / inventory", 12000),
                       ("Direct labour costs", 8000),
                       ("Manufacturing overhead", 3000)]:
        label_row(ws, r, label)
        input_cell(ws, r, 5, val)
        pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})", color=GRAY_MID)
        cogs_rows.append(r)
        r += 1

    cogs_total_row = r
    total_row(ws, r, "Total COGS",
              f"=SUM(E{cogs_rows[0]}:E{cogs_rows[-1]})", col_val=5,
              bg=AMBER_LIGHT, txt_color=AMBER)
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})",
                bold=True, color=AMBER, bg=AMBER_LIGHT)
    r += 1

    gp_row = r
    total_row(ws, r, "GROSS PROFIT",
              f"=E{net_rev_row}-E{cogs_total_row}", col_val=5)
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})",
                bold=True, color=GREEN_MID, bg=GREEN_LIGHT)
    r += 2

    # ── OPERATING EXPENSES ───────────────────────────────────────────────
    section_header(ws, r, "  OPERATING EXPENSES", 2, 6)
    r += 1
    opex_rows = []
    for label, val in [("Rent & utilities", 4500),
                       ("Staff salaries", 15000),
                       ("Marketing & advertising", 3000),
                       ("Software & subscriptions", 1200),
                       ("Insurance", 800),
                       ("Admin & miscellaneous", 2200)]:
        label_row(ws, r, label)
        input_cell(ws, r, 5, val)
        pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})", color=GRAY_MID)
        opex_rows.append(r)
        r += 1

    opex_total_row = r
    total_row(ws, r, "Total Operating Expenses",
              f"=SUM(E{opex_rows[0]}:E{opex_rows[-1]})", col_val=5,
              bg=AMBER_LIGHT, txt_color=AMBER)
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})",
                bold=True, color=AMBER, bg=AMBER_LIGHT)
    r += 1

    ebit_row = r
    total_row(ws, r, "EBIT (Operating Profit)",
              f"=E{gp_row}-E{opex_total_row}", col_val=5)
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})",
                bold=True, color=GREEN_MID, bg=GREEN_LIGHT)
    r += 2

    # ── BELOW THE LINE ───────────────────────────────────────────────────
    section_header(ws, r, "  BELOW THE LINE", 2, 6)
    r += 1

    int_row = r
    label_row(ws, r, "Interest expense")
    input_cell(ws, r, 5, 800)
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})", color=GRAY_MID)
    r += 1

    ebt_row = r
    total_row(ws, r, "EBT (Profit Before Tax)",
              f"=E{ebit_row}-E{int_row}", col_val=5)
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})",
                bold=True, color=GREEN_MID, bg=GREEN_LIGHT)
    r += 1

    tax_rate_row = r
    label_row(ws, r, "Tax rate (%)")
    c = input_cell(ws, r, 5, 0.25, num_fmt='0%')
    r += 1

    tax_row = r
    label_row(ws, r, "Income tax")
    formula_cell(ws, r, 5, f"=IF(E{ebt_row}>0,E{ebt_row}*E{tax_rate_row},0)")
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})", color=GRAY_MID)
    r += 1

    np_row = r
    ws.row_dimensions[r].height = 26
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(row=r, column=2, value="NET PROFIT (BOTTOM LINE)")
    c.font = font(bold=True, size=13, color=WHITE)
    c.fill = fill(DARK_BG)
    c.alignment = align(h="left")
    ws.cell(row=r, column=3).fill = fill(DARK_BG)
    ws.cell(row=r, column=4).fill = fill(DARK_BG)
    formula_cell(ws, r, 5, f"=E{ebt_row}-E{tax_row}",
                 bold=True, color=WHITE, bg=DARK_BG,
                 border=thick_bottom(GREEN), size=13)
    pct_formula(ws, r, 6, f"=IF(E{net_rev_row}=0,0,E{r}/E{net_rev_row})",
                bold=True, color=GREEN, bg=DARK_BG)
    r += 2

    # ── KPI SUMMARY ──────────────────────────────────────────────────────
    section_header(ws, r, "  KEY METRICS SUMMARY", 2, 6)
    r += 1
    kpis = [
        ("Gross Profit Margin", f"=IF(E{net_rev_row}=0,0,E{gp_row}/E{net_rev_row})"),
        ("Operating Profit Margin (EBIT)", f"=IF(E{net_rev_row}=0,0,E{ebit_row}/E{net_rev_row})"),
        ("Net Profit Margin", f"=IF(E{net_rev_row}=0,0,E{np_row}/E{net_rev_row})"),
        ("COGS as % of Revenue", f"=IF(E{net_rev_row}=0,0,E{cogs_total_row}/E{net_rev_row})"),
        ("OpEx as % of Revenue", f"=IF(E{net_rev_row}=0,0,E{opex_total_row}/E{net_rev_row})"),
    ]
    for label, formula in kpis:
        label_row(ws, r, label)
        pct_formula(ws, r, 5, formula, color=BLUE)
        r += 1

    r += 1
    note(ws, r, 2, "INSTRUCTIONS: Enter your numbers in BLUE cells only. "
         "All other cells calculate automatically. Blue = input. Black = formula.")

    # freeze panes
    ws.freeze_panes = "B5"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — BALANCE SHEET
# ═══════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [1, 34, 4, 4, 16, 16, 4])

    ws.merge_cells("B1:F1")
    c = ws["B1"]
    c.value = "BALANCE SHEET"
    c.font = font(bold=True, size=20, color=WHITE, name="Arial")
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "The Solo Operator's Financial Toolkit  •  harianand07"
    c.font = font(size=10, color=GRAY_MID, italic=True)
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 22
    for col, label in [(2,"Line Item"), (5,"Amount ($)"), (6,"Notes / Ratio")]:
        c = ws.cell(row=4, column=col, value=label)
        c.font = font(bold=True, size=10, color=WHITE)
        c.fill = fill(GREEN_MID)
        c.alignment = align(h="center" if col>2 else "left")
    ws.row_dimensions[5].height = 6

    r = 6

    # ── ASSETS ───────────────────────────────────────────────────────────
    section_header(ws, r, "  ASSETS", 2, 6)
    r += 1
    sub_header(ws, r, "Current Assets", 2, 6, bg="FF1D9E75", txt=WHITE)
    r += 1

    ca_rows = []
    for label, val in [("Cash & bank balance", 22000),
                       ("Accounts receivable", 14000),
                       ("Inventory / stock", 18000),
                       ("Prepaid expenses", 3000),
                       ("Short-term investments", 5000)]:
        label_row(ws, r, label, indent=True)
        input_cell(ws, r, 5, val)
        ca_rows.append(r)
        r += 1

    ca_total_row = r
    total_row(ws, r, "Total Current Assets",
              f"=SUM(E{ca_rows[0]}:E{ca_rows[-1]})", col_val=5)
    r += 1

    sub_header(ws, r, "Non-Current Assets (Long-Term)", 2, 6, bg="FF1D9E75", txt=WHITE)
    r += 1
    nca_rows = []
    equip_row = r
    for label, val in [("Property & equipment (cost)", 45000),
                       ("Less: accumulated depreciation", 8000),
                       ("Intangible assets / goodwill", 5000),
                       ("Long-term investments", 10000)]:
        label_row(ws, r, label, indent=True)
        input_cell(ws, r, 5, val)
        nca_rows.append(r)
        r += 1

    nca_total_row = r
    formula_cell(ws, r, 5,
                 f"=E{nca_rows[0]}-E{nca_rows[1]}+E{nca_rows[2]}+E{nca_rows[3]}",
                 bold=True, color=GREEN_MID, bg=GREEN_LIGHT,
                 border=thick_bottom(GREEN))
    ws.cell(row=r, column=2, value="Net Non-Current Assets").font = font(bold=True, color=GREEN_MID)
    ws.cell(row=r, column=2).fill = fill(GREEN_LIGHT)
    ws.row_dimensions[r].height = 20
    r += 1

    ta_row = r
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(row=r, column=2, value="TOTAL ASSETS")
    c.font = font(bold=True, size=13, color=WHITE)
    c.fill = fill(DARK_BG)
    c.alignment = align(h="left")
    for col in [3,4]:
        ws.cell(row=r, column=col).fill = fill(DARK_BG)
    formula_cell(ws, r, 5, f"=E{ca_total_row}+E{nca_total_row}",
                 bold=True, color=WHITE, bg=DARK_BG,
                 border=thick_bottom(GREEN), size=13)
    r += 2

    # ── LIABILITIES ──────────────────────────────────────────────────────
    section_header(ws, r, "  LIABILITIES", 2, 6)
    r += 1
    sub_header(ws, r, "Current Liabilities (Due within 1 year)", 2, 6,
               bg="FFB91C1C", txt=WHITE)
    r += 1

    cl_rows = []
    for label, val in [("Accounts payable", 9000),
                       ("Short-term loans", 6000),
                       ("Accrued expenses", 4000),
                       ("Taxes payable", 2500),
                       ("Unearned revenue", 1500)]:
        label_row(ws, r, label, indent=True)
        input_cell(ws, r, 5, val)
        cl_rows.append(r)
        r += 1

    cl_total_row = r
    total_row(ws, r, "Total Current Liabilities",
              f"=SUM(E{cl_rows[0]}:E{cl_rows[-1]})", col_val=5,
              bg=RED_LIGHT, txt_color=RED)
    r += 1

    sub_header(ws, r, "Non-Current Liabilities (Long-Term)", 2, 6,
               bg="FFB91C1C", txt=WHITE)
    r += 1
    ncl_rows = []
    for label, val in [("Long-term bank loans", 20000),
                       ("Deferred tax liability", 2500),
                       ("Other long-term liabilities", 0)]:
        label_row(ws, r, label, indent=True)
        input_cell(ws, r, 5, val)
        ncl_rows.append(r)
        r += 1

    ncl_total_row = r
    total_row(ws, r, "Total Non-Current Liabilities",
              f"=SUM(E{ncl_rows[0]}:E{ncl_rows[-1]})", col_val=5,
              bg=RED_LIGHT, txt_color=RED)
    r += 1

    tl_row = r
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(row=r, column=2, value="TOTAL LIABILITIES")
    c.font = font(bold=True, size=13, color=WHITE)
    c.fill = fill("FF7F1D1D")
    c.alignment = align(h="left")
    for col in [3,4]:
        ws.cell(row=r, column=col).fill = fill("FF7F1D1D")
    formula_cell(ws, r, 5, f"=E{cl_total_row}+E{ncl_total_row}",
                 bold=True, color=WHITE, bg="FF7F1D1D",
                 border=thick_bottom(RED), size=13)
    r += 2

    # ── EQUITY ───────────────────────────────────────────────────────────
    section_header(ws, r, "  OWNER'S EQUITY", 2, 6)
    r += 1

    eq_rows = []
    for label, val in [("Owner's capital contributed", 30000),
                       ("Retained earnings (prior years)", 15000),
                       ("Current year net profit", 0)]:
        label_row(ws, r, label, indent=True)
        input_cell(ws, r, 5, val)
        eq_rows.append(r)
        r += 1

    note(ws, r, 2, "Tip: Link 'Current year net profit' from your Income Statement.")
    r += 1

    eq_total_row = r
    formula_cell(ws, r, 5, f"=E{ta_row}-E{tl_row}",
                 bold=True, color=WHITE, bg=DARK_BG,
                 border=thick_bottom(GREEN), size=13)
    ws.row_dimensions[r].height = 24
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(row=r, column=2, value="TOTAL OWNER'S EQUITY  (Assets − Liabilities)")
    c.font = font(bold=True, size=12, color=WHITE)
    c.fill = fill(DARK_BG)
    c.alignment = align(h="left")
    for col in [3,4]:
        ws.cell(row=r, column=col).fill = fill(DARK_BG)
    r += 2

    # ── FINANCIAL RATIOS ─────────────────────────────────────────────────
    section_header(ws, r, "  FINANCIAL HEALTH RATIOS", 2, 6)
    r += 1

    ratios = [
        ("Current Ratio  (ideal > 1.5)", f"=IFERROR(E{ca_total_row}/E{cl_total_row},0)", '0.00x', "Measures short-term liquidity"),
        ("Quick Ratio  (ideal > 1.0)", f"=IFERROR((E{ca_rows[0]}+E{ca_rows[1]})/E{cl_total_row},0)", '0.00x', "Cash + receivables vs current liabilities"),
        ("Debt Ratio  (ideal < 0.5)", f"=IFERROR(E{tl_row}/E{ta_row},0)", '0.0%', "% of assets financed by debt"),
        ("Debt-to-Equity  (ideal < 1.0)", f"=IFERROR(E{tl_row}/E{eq_total_row},0)", '0.00x', "Debt vs owner equity"),
        ("Working Capital ($)", f"=E{ca_total_row}-E{cl_total_row}", '$#,##0;($#,##0);"-"', "Current assets minus current liabilities"),
        ("Balance Check (must = 0)", f"=E{ta_row}-E{tl_row}-E{eq_total_row}", '$#,##0;($#,##0);"-"', "Must be zero — if not, recheck inputs"),
    ]
    for label, formula, fmt, tip in ratios:
        label_row(ws, r, label)
        formula_cell(ws, r, 5, formula, color=BLUE, num_fmt=fmt)
        note(ws, r, 6, tip)
        r += 1

    r += 1
    note(ws, r, 2, "INSTRUCTIONS: Enter your numbers in BLUE cells. All ratios calculate automatically. "
         "The Balance Check must equal $0 — if it doesn't, check your equity inputs.")

    ws.freeze_panes = "B5"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — CASH FLOW STATEMENT
# ═══════════════════════════════════════════════════════════════════════════
def build_cash_flow(wb):
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [1, 30, 4, 4, 14, 14, 14, 14, 14, 14, 4])

    ws.merge_cells("B1:J1")
    c = ws["B1"]
    c.value = "CASH FLOW STATEMENT  —  12-MONTH TRACKER"
    c.font = font(bold=True, size=18, color=WHITE, name="Arial")
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("B2:J2")
    c = ws["B2"]
    c.value = "The Solo Operator's Financial Toolkit  •  harianand07"
    c.font = font(size=10, color=GRAY_MID, italic=True)
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 22

    months = ["Jan","Feb","Mar","Apr","May","Jun"]
    cols   = [5,6,7,8,9,10]

    for col, month in zip(cols, months):
        c = ws.cell(row=4, column=col, value=month)
        c.font = font(bold=True, size=10, color=WHITE)
        c.fill = fill(GREEN_MID)
        c.alignment = align(h="center")

    c = ws.cell(row=4, column=2, value="Line Item")
    c.font = font(bold=True, size=10, color=WHITE)
    c.fill = fill(GREEN_MID)
    ws.row_dimensions[5].height = 6

    r = 6

    # ── OPENING BALANCE ──────────────────────────────────────────────────
    section_header(ws, r, "  OPENING CASH BALANCE", 2, 10)
    r += 1

    open_row = r
    label_row(ws, r, "Cash at start of month")
    input_cell(ws, r, 5, 15000)   # Jan opening — user fills
    for col in cols[1:]:
        # Feb onward = prior month closing
        formula_cell(ws, r, col, f"={get_column_letter(col-1)}{r+14}")
    r += 1
    ws.row_dimensions[r].height = 6
    r += 1

    # ── OPERATING INFLOWS ────────────────────────────────────────────────
    section_header(ws, r, "  OPERATING ACTIVITIES — CASH IN", 2, 10)
    r += 1
    in_items = [
        ("Cash received from customers", [38000,41000,37000,44000,42000,46000]),
        ("Other operating receipts",     [1000,1200,800,1500,1100,1300]),
    ]
    in_rows = []
    for label, vals in in_items:
        label_row(ws, r, label, indent=True)
        for col, val in zip(cols, vals):
            input_cell(ws, r, col, val)
        in_rows.append(r)
        r += 1

    total_in_row = r
    label_row(ws, r, "Total Cash Inflows")
    for col in cols:
        formula_cell(ws, r, col, f"=SUM({get_column_letter(col)}{in_rows[0]}:{get_column_letter(col)}{in_rows[-1]})",
                     bold=True, color=GREEN_MID, bg=GREEN_LIGHT)
    r += 1
    ws.row_dimensions[r].height = 6
    r += 1

    # ── OPERATING OUTFLOWS ───────────────────────────────────────────────
    section_header(ws, r, "  OPERATING ACTIVITIES — CASH OUT", 2, 10)
    r += 1
    out_items = [
        ("Payments to suppliers",     [14000,15000,13500,16000,14500,15500]),
        ("Staff salaries paid",        [12000,12000,12000,12000,12000,12000]),
        ("Rent & utilities",           [4000,4000,4000,4000,4000,4000]),
        ("Marketing & advertising",    [3000,2500,3500,4000,3000,3500]),
        ("Tax payments",               [2500,0,2500,0,2500,0]),
        ("Other operating payments",   [1800,2000,1600,2200,1900,2100]),
    ]
    out_rows = []
    for label, vals in out_items:
        label_row(ws, r, label, indent=True)
        for col, val in zip(cols, vals):
            input_cell(ws, r, col, val)
        out_rows.append(r)
        r += 1

    total_out_row = r
    label_row(ws, r, "Total Cash Outflows")
    for col in cols:
        formula_cell(ws, r, col, f"=SUM({get_column_letter(col)}{out_rows[0]}:{get_column_letter(col)}{out_rows[-1]})",
                     bold=True, color=RED, bg=RED_LIGHT)
    r += 1

    ocf_row = r
    label_row(ws, r, "NET OPERATING CASH FLOW")
    for col in cols:
        formula_cell(ws, r, col,
                     f"={get_column_letter(col)}{total_in_row}-{get_column_letter(col)}{total_out_row}",
                     bold=True, color=WHITE, bg=GREEN_MID)
    ws.row_dimensions[r].height = 22
    r += 2

    # ── INVESTING ────────────────────────────────────────────────────────
    section_header(ws, r, "  INVESTING ACTIVITIES", 2, 10)
    r += 1
    inv_items = [
        ("Equipment / asset purchases (outflow)",  [5000,0,0,8000,0,0]),
        ("Asset sales / disposals (inflow)",       [0,1200,0,0,2000,0]),
    ]
    inv_rows = []
    for label, vals in inv_items:
        label_row(ws, r, label, indent=True)
        for col, val in zip(cols, vals):
            input_cell(ws, r, col, val)
        inv_rows.append(r)
        r += 1

    inv_net_row = r
    label_row(ws, r, "Net Investing Cash Flow")
    for col in cols:
        formula_cell(ws, r, col,
                     f"={get_column_letter(col)}{inv_rows[1]}-{get_column_letter(col)}{inv_rows[0]}",
                     bold=True, color=AMBER, bg=AMBER_LIGHT)
    r += 2

    # ── FINANCING ────────────────────────────────────────────────────────
    section_header(ws, r, "  FINANCING ACTIVITIES", 2, 10)
    r += 1
    fin_items = [
        ("Loans received (inflow)",    [0,0,10000,0,0,0]),
        ("Loan repayments (outflow)",  [2000,2000,2000,2000,2000,2000]),
        ("Owner drawings / dividends", [3000,3000,3000,3000,3000,3000]),
    ]
    fin_rows = []
    for label, vals in fin_items:
        label_row(ws, r, label, indent=True)
        for col, val in zip(cols, vals):
            input_cell(ws, r, col, val)
        fin_rows.append(r)
        r += 1

    fin_net_row = r
    label_row(ws, r, "Net Financing Cash Flow")
    for col in cols:
        formula_cell(ws, r, col,
                     f"={get_column_letter(col)}{fin_rows[0]}-{get_column_letter(col)}{fin_rows[1]}-{get_column_letter(col)}{fin_rows[2]}",
                     bold=True, color=BLUE, bg=BLUE_LIGHT)
    r += 2

    # ── NET & CLOSING ────────────────────────────────────────────────────
    section_header(ws, r, "  CLOSING BALANCE", 2, 10, bg="FF111827")
    r += 1

    net_row = r
    label_row(ws, r, "Net Change in Cash")
    for col in cols:
        formula_cell(ws, r, col,
                     f"={get_column_letter(col)}{ocf_row}+{get_column_letter(col)}{inv_net_row}+{get_column_letter(col)}{fin_net_row}",
                     bold=True, color=FORMULA_BLK)
    r += 1

    close_row = r
    ws.row_dimensions[r].height = 26
    c = ws.cell(row=r, column=2, value="CLOSING CASH BALANCE")
    c.font = font(bold=True, size=13, color=WHITE)
    c.fill = fill(DARK_BG)
    c.alignment = align(h="left")
    for col in cols:
        formula_cell(ws, r, col,
                     f"={get_column_letter(col)}{open_row}+{get_column_letter(col)}{net_row}",
                     bold=True, color=GREEN, bg=DARK_BG,
                     border=thick_bottom(GREEN), size=13)
    r += 2

    # ── RUNWAY & KPI SUMMARY ─────────────────────────────────────────────
    section_header(ws, r, "  CASH RUNWAY & KEY METRICS", 2, 10)
    r += 1

    for col in cols:
        label_row(ws, r, "Monthly cash burn rate")
        formula_cell(ws, r, col,
                     f"={get_column_letter(col)}{total_out_row}+{get_column_letter(col)}{fin_rows[1]}+{get_column_letter(col)}{fin_rows[2]}",
                     color=AMBER)
    burn_row = r
    r += 1

    for col in cols:
        label_row(ws, r, "Runway (months remaining)")
        formula_cell(ws, r, col,
                     f"=IFERROR({get_column_letter(col)}{close_row}/{get_column_letter(col)}{burn_row},99)",
                     color=BLUE, num_fmt="0.0")
    r += 1

    for col in cols:
        label_row(ws, r, "Operating cash flow margin")
        formula_cell(ws, r, col,
                     f"=IFERROR({get_column_letter(col)}{ocf_row}/{get_column_letter(col)}{total_in_row},0)",
                     color=GREEN_MID, num_fmt="0.0%")
    r += 2

    note(ws, r, 2, "INSTRUCTIONS: Blue cells = your inputs. Enter monthly figures for each column. "
         "Feb-Jun opening balances auto-link from prior month closing. "
         "Runway shows how many months of cash you have left at current burn rate.")

    ws.freeze_panes = "E5"
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# INSTRUCTIONS SHEET
# ═══════════════════════════════════════════════════════════════════════════
def build_instructions(wb):
    ws = wb.create_sheet("START HERE")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 2

    ws.merge_cells("B1:B1")
    ws.row_dimensions[1].height = 50
    c = ws["B1"]
    c.value = "THE SOLO OPERATOR'S FINANCIAL TOOLKIT"
    c.font = font(bold=True, size=22, color=WHITE, name="Arial")
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")

    ws.row_dimensions[2].height = 20
    c = ws["B2"]
    c.value = "Built by Hari Anand  •  linkedin.com/in/harianand07"
    c.font = font(size=11, color=GRAY_MID, italic=True)
    c.fill = fill(DARK_BG)
    c.alignment = align(h="center")

    ws.row_dimensions[3].height = 12

    steps = [
        ("STEP 1 — Income Statement", "Go to the 'Income Statement' tab. Enter your revenue, COGS, and expenses in the BLUE cells. Net profit calculates automatically."),
        ("STEP 2 — Balance Sheet", "Go to the 'Balance Sheet' tab. Enter your assets, liabilities, and equity in BLUE cells. The Balance Check at the bottom must equal $0."),
        ("STEP 3 — Cash Flow", "Go to the 'Cash Flow' tab. Enter monthly figures for 6 months. Opening balances auto-link month to month. Watch your runway metric."),
        ("COLOUR CODE", "BLUE text = input cells (you fill these)\nBLACK text = formula cells (auto-calculated, do not edit)\nGREEN = healthy / profit\nRED = expense / loss / danger\nAMBER = warning / expense category"),
        ("HOW TO USE", "1. Fill in your real business numbers each month\n2. Share the Income Statement with your accountant\n3. Use the Balance Sheet when applying for a bank loan\n4. Monitor Cash Flow every week to avoid running out of cash"),
        ("SUPPORT / CUSTOM WORK", "Need a custom version for your business? Want dashboards, charts, or AI integration?\nConnect: linkedin.com/in/harianand07"),
    ]

    r = 4
    for title, body in steps:
        ws.row_dimensions[r].height = 24
        c = ws.cell(row=r, column=2, value=title)
        c.font = font(bold=True, size=12, color=WHITE)
        c.fill = fill(GREEN_MID)
        c.alignment = align(h="left")
        r += 1

        for line in body.split('\n'):
            ws.row_dimensions[r].height = 18
            c = ws.cell(row=r, column=2, value=line)
            c.font = font(size=11, color=GRAY_DARK)
            c.alignment = align(h="left", wrap=True)
            r += 1

        ws.row_dimensions[r].height = 8
        r += 1

    return ws


# ═══════════════════════════════════════════════════════════════════════════
# ASSEMBLE WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)   # remove default Sheet

build_instructions(wb)
build_income_statement(wb)
build_balance_sheet(wb)
build_cash_flow(wb)

# Set tab colours
wb["START HERE"].sheet_properties.tabColor        = "1DBF73"
wb["Income Statement"].sheet_properties.tabColor  = "0F6E56"
wb["Balance Sheet"].sheet_properties.tabColor     = "1D4ED8"
wb["Cash Flow"].sheet_properties.tabColor         = "D97706"

out = "Financial_Toolkit_HariAnand.xlsx"
wb.save(out)
print("Saved:", out)